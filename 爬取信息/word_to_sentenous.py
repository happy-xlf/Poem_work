#抽取诗词中包含飞花令词的句子：保存在sentences文件夹下

import pandas as pd
import xlwt
import openpyxl

#读取飞花令
def read_word():
    data=pd.read_excel('data2/word.xlsx')
    words=data.word
    return words

#遍历诗句
def read(file,words,write_file):
    data=pd.read_excel(file)
    title=data.title
    content=data.content
    author=data.author
    #进行切分出单句
    ans_sentens = []
    ans_author = []
    ans_title = []
    ans_key = []
    for i in range(len(title)):
        print("第"+str(i)+"个")
        cont=content[i]
        aut=author[i]
        tit=title[i]
        sents=cont.replace('\n','').split('。')
        for it in sents:
            key_list = []
            for k in words:
                if it.find(k)!=-1:
                    key_list.append(k)
            if len(key_list)!=0:
                ans_sentens.append(it)
                ans_author.append(aut)
                ans_title.append(tit)
                ans_key.append(",".join(key_list))

    #存储对应的key，author，title，sentenous
    xl = openpyxl.Workbook()
    # 调用对象的add_sheet方法
    sheet1 = xl.create_sheet(index=0)
    sheet1.cell(1, 1, "sentens")
    sheet1.cell(1, 2, "author")
    sheet1.cell(1, 3, "title")
    sheet1.cell(1, 4, "keys")

    for i in range(0, len(ans_key)):
        sheet1.cell(i + 2, 1, ans_sentens[i])
        sheet1.cell(i + 2, 2, ans_author[i])
        sheet1.cell(i + 2, 3, ans_title[i])
        sheet1.cell(i + 2, 4, ans_key[i])
    xl.save(write_file)
    print("保存成功到-"+write_file)

#获取指定文件夹下的excel
import os
def get_filename(path,filetype):  # 输入路径、文件类型例如'.xlsx'
    name = []
    for root,dirs,files in os.walk(path):
        for i in files:
            if os.path.splitext(i)[1]==filetype:
                name.append(i)
    return name            # 输出由有后缀的文件名组成的列表

if __name__ == '__main__':
    file='data/'
    words=read_word()
    list = get_filename(file, '.xlsx')
    for i in range(len(list)):
        new_file=file+list[i]
        print(new_file)
        sentences_file = "sentences/sentence" + str(i+1) + ".xlsx"
        read(new_file,words,sentences_file)