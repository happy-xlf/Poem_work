import pandas as pd
from 情感分析.词相似度.utils.mysqlhelper import MySqLHelper
import openpyxl

def emotion():
    data=pd.read_excel('new_emotion.xlsx')
    similar=data.get('similar')
    value = data.get('val')
    sad_list=str(similar[0]).split(',')
    fear_list=str(similar[1]).split(',')
    happy_list=str(similar[2]).split(',')
    anger_list=str(similar[3]).split(',')
    think_list=str(similar[4]).split(',')
    like_list=str(similar[5]).split(',')
    worry_list=str(similar[6]).split(',')

    sad_val_list = str(value[0]).split(',')
    fear_val_list = str(value[1]).split(',')
    happy_val_list = str(value[2]).split(',')
    anger_val_list = str(value[3]).split(',')
    think_val_list = str(value[4]).split(',')
    like_val_list = str(value[5]).split(',')
    worry_val_list = str(value[6]).split(',')
    return sad_list,fear_list,happy_list,anger_list,think_list,like_list,worry_list,sad_val_list,fear_val_list,happy_val_list,anger_val_list,think_val_list,like_val_list,worry_val_list

def test_sentence(sentence):
    sad_list, fear_list, happy_list, anger_list, think_list, like_list, worry_list,sad_val_list,fear_val_list,happy_val_list,anger_val_list,think_val_list,like_val_list,worry_val_list=emotion()
    sad=fear=happy=anger=think=like=worry=0
    for k in sentence:
        if k in sad_list:
            sad+=float(sad_val_list[sad_list.index(k)])
        elif k in fear_list:
            fear+=float(fear_val_list[fear_list.index(k)])
        elif k in happy_list:
            happy+=float(happy_val_list[happy_list.index(k)])
        elif k in anger_list:
            anger+=float(anger_val_list[anger_list.index(k)])
        elif k in think_list:
            think+=float(think_val_list[think_list.index(k)])
        elif k in like_list:
            like+=float(like_val_list[like_list.index(k)])
        elif k in worry_list:
            worry+=float(worry_val_list[worry_list.index(k)])
    ans=max(sad,fear,happy,anger,think,like,worry)
    scord_list=[]
    scord_list.append(sad)
    scord_list.append(fear)
    scord_list.append(happy)
    scord_list.append(anger)
    scord_list.append(think)
    scord_list.append(like)
    scord_list.append(worry)
    emotion_list=['悲','惧','乐','怒','思','喜','忧']
    i=0
    for i in range(len(scord_list)):
        if scord_list[i]==ans:
            #print(emotion_list[i])
            break
    if ans!=0:
        return emotion_list[i],ans
    else:
        return '无',0

def read():
    db=MySqLHelper()
    sql="select * from qing"
    ret,count=db.selectall(sql=sql)
    content_list=[]
    for row in ret:
        content=str(row[3]).replace('\n','').replace('□','')
        content_list.append(content)
    ans_emotion_content=[]
    print(len(content_list))
    for i in range(len(content_list)):
        print('第'+str(i)+'个')
        content=content_list[i]
        #print(content)
        ans_content=[]
        content_l=str(content).split('。')
        for k in content_l:
            kk=str(k).split('，')
            for it in kk:
                if it!='':
                    ans_content.append(it)
        ans_emotion = {}
        for sentence in ans_content:
            #print(sentence)
            emot,score=test_sentence(sentence)
            if emot!='无':
                #print(emot+str(score))
                if emot not in ans_emotion.keys():
                    ans_emotion[emot] = score
                else:
                    ans_emotion[emot]+=score
        #print(sorted(ans_emotion.items(), key=lambda item: item[1], reverse=True))
        if len(ans_emotion.items())==0:
            #print('整篇文章情感：无')
            ans_emotion_content.append('无')
        else:
            ans_emotion=dict(sorted(ans_emotion.items(), key=lambda item: item[1], reverse=True))
            for key,value in ans_emotion.items():
                #print('整篇文章情感：'+key)
                ans_emotion_content.append(key)
                break
    # import xlwt
    #
    # xl = xlwt.Workbook()
    # # 调用对象的add_sheet方法
    # sheet1 = xl.add_sheet('sheet1', cell_overwrite_ok=True)
    #
    # sheet1.write(0, 0, "sentence")
    # sheet1.write(0, 1, 'label')
    # print(len(content_list))
    # for i in range(0, len(content_list)):
    #     sheet1.write(i + 1, 0, content_list[i].replace('\n',''))
    #     sheet1.write(i + 1, 1, ans_emotion_content[i])
    # xl.save("train_song.xlsx")

    write_file='train_qing.xlsx'
    xl = openpyxl.Workbook()
    # 调用对象的add_sheet方法
    sheet1 = xl.create_sheet(index=0)
    sheet1.cell(1, 1, "sentence")
    sheet1.cell(1, 2, "label")

    for i in range(0, len(content_list)):
        sheet1.cell(i + 2, 1, content_list[i])
        sheet1.cell(i + 2, 2, ans_emotion_content[i])
    xl.save(write_file)
    print("保存成功到-"+write_file)



if __name__ == '__main__':
    read()