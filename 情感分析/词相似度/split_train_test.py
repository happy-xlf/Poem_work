# -*- coding: utf-8 -*- 
# @Time : 2021/12/25 8:55 
# @Author : xlf
# @File : split_train_test.py
import pandas as pd


def read_excel():
    excel_list=['train_tang','train_song','train_yuan','train_ming','train_qing']
    sentence_list=[]
    label_list=[]
    for it in excel_list:
        data=pd.read_excel(it+".xlsx")
        sentence=data.get('sentence')
        label=data.get('label')
        sentence_list.append(sentence)
        label_list.append(label)
    return sentence_list,label_list

def count_list():
    sentence_list,label_list=read_excel()
    ans_sentence=[]
    ans_label=[]
    for i in range(len(sentence_list)):
        sentence=sentence_list[i]
        label=label_list[i]
        for lb in label:
            ans_label.append(lb)
        for se in sentence:
            ans_sentence.append(se)
    return ans_sentence,ans_label

#2100,7500,12000,24000,27000,28500,36000
#{'惧': 7204, '怒': 25736, '忧': 43920, '喜': 79206, '思': 90851, '乐': 95283, '悲': 125214}


def split_data():
    sentence_list,label_list=count_list()
    emotion = ['悲', '惧', '乐', '怒', '思', '喜', '忧']
    ans={}
    train_senten=[]
    train_label=[]
    test_senten=[]
    test_label=[]
    k1=k2=k3=k4=k5=k6=k7=0
    for i in range(len(sentence_list)):
        lab=label_list[i]
        sente=sentence_list[i]
        if lab in emotion:
            if lab == '悲':
                if k1<=36000:
                    k1+=1
                    test_senten.append(sente)
                    test_label.append(0)
                else:
                    train_label.append(0)
                    train_senten.append(sente)
            elif lab == '惧':
                if k2<=2100:
                    k2+=1
                    test_senten.append(sente)
                    test_label.append(1)
                else:
                    train_label.append(1)
                    train_senten.append(sente)
            elif lab == '乐':
                if k3<=28500:
                    k3+=1
                    test_senten.append(sente)
                    test_label.append(2)
                else:
                    train_label.append(2)
                    train_senten.append(sente)
            elif lab == '怒':
                if k4<=7500:
                    k4+=1
                    test_senten.append(sente)
                    test_label.append(3)
                else:
                    train_label.append(3)
                    train_senten.append(sente)
            elif lab == '思':
                if k5<=27000:
                    k5+=1
                    test_senten.append(sente)
                    test_label.append(4)
                else:
                    train_label.append(4)
                    train_senten.append(sente)
            elif lab == '喜':
                if k6<=24000:
                    k6+=1
                    test_senten.append(sente)
                    test_label.append(5)
                else:
                    train_label.append(5)
                    train_senten.append(sente)
            elif lab == '忧':
                if k7<=12000:
                    k7+=1
                    test_senten.append(sente)
                    test_label.append(6)
                else:
                    train_label.append(6)
                    train_senten.append(sente)
    # import xlwt
    #
    # xl = xlwt.Workbook()
    # # 调用对象的add_sheet方法
    # sheet1 = xl.add_sheet('sheet1', cell_overwrite_ok=True)
    #
    # sheet1.write(0, 0, "sentence")
    # sheet1.write(0, 1, 'label')
    # for i in range(0, len(test_senten)):
    #     sheet1.write(i + 1, 0, test_senten[i])
    #     sheet1.write(i + 1, 1, test_label[i])
    # xl.save("data/test.xlsx")

    import openpyxl
    write_file = 'data/train.xlsx'
    xl = openpyxl.Workbook()
    # 调用对象的add_sheet方法
    sheet1 = xl.create_sheet(index=0)
    sheet1.cell(1, 1, "sentence")
    sheet1.cell(1, 2, "label")

    for i in range(0, len(train_senten)):
        sheet1.cell(i + 2, 1, train_senten[i])
        sheet1.cell(i + 2, 2, train_label[i])
    xl.save(write_file)
    print("保存成功到-" + write_file)


if __name__ == '__main__':
    split_data()